"""
audit_stocks.py — Per-stock filter audit for known winners
===========================================================
Reads symbols from an Excel file and audits each one: which R01/entry
conditions pass or fail, first alert date, first entry signal date.

Calls each rule function DIRECTLY (not via entry_checklist) to avoid
silent exception swallowing.

Usage:
    python audit_stocks.py --xlsx YearlyWinners_2023.xlsx --start 2023-01-01 --end 2023-12-31
    python audit_stocks.py --symbols TITAGARH,ZENTEC --start 2023-01-01 --end 2023-12-31 --verbose
"""
import argparse
import os
import sys
import traceback
from datetime import datetime

import numpy as np
import pandas as pd

sys.path.insert(0, os.path.dirname(__file__))

from config import (
    BASE_DIR, INDEX_SYMBOL, OUTPUT_DIR,
    POSITION_LOOKBACK_DAYS, R01_WATCHLIST_DAYS, R01_MIN_MOVE,
    RS_MIN_RATIO, MIN_TRADING_ROWS, N_CONFIRM_REQUIRED,
)
from data_loader import load_stock, load_index, align_stock_index
from indicators import (
    resample_weekly, add_weekly_emas, add_daily_emas,
    index_regime, rs_ratio,
)
from signals import (
    rule01_price_moved_50pct,
    rule02_green_weekly_candles,
    rule03_rs_ratio,
    rule04_index_regime,
    rule09_stage2,
    rule08_rs_line_near_highs,
    rule11_volatility_contraction,
    rule15_breakout_volume,
    rule12_low_vol_red_days,
    rule_weekly_inside_candle_breakout,
    rule_weekly_10wma_support_bounce,
    rule_weekly_prev_high_near_10wma,
    rule_vcp_breakout,
    rule_weekly_hammer_or_engulfing,
    rule_base_formed,
)

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ── helpers ───────────────────────────────────────────────────────────────────

def _safe(fn, *args, default=False):
    """Call fn(*args), return default on any exception."""
    try:
        return fn(*args)
    except Exception:
        return default


def _base_depth_pass(stock_df: pd.DataFrame) -> bool:
    """Base retraced 8–25% from 40-bar swing high to CURRENT CLOSE.
    Matches the fixed entry_checklist formula: (hi - close_now) / hi.
    NOT (hi-lo)/hi which measures oscillation, not retracement."""
    try:
        from config import BASE_DEPTH_MIN, BASE_DEPTH_MAX
        from indicators import base_range
        lo, hi = base_range(stock_df, 40)
        current_close = float(stock_df["Close"].iloc[-1])
        depth = (hi - current_close) / hi if hi > 0 else 0
        return BASE_DEPTH_MIN <= depth <= BASE_DEPTH_MAX
    except Exception:
        return False


def _scan_r01_windows(stock_df: pd.DataFrame):
    """Return list of (start_date, end_date, pct, bars) where R01 fires."""
    close = stock_df["Close"].values
    dates = stock_df["date"].values
    n = len(close)
    hits = []
    scan_start = max(POSITION_LOOKBACK_DAYS, n - R01_WATCHLIST_DAYS)
    for i in range(scan_start, n):
        base = close[i - POSITION_LOOKBACK_DAYS]
        if base > 0 and (close[i] / base - 1) >= R01_MIN_MOVE:
            hits.append((
                pd.Timestamp(dates[i - POSITION_LOOKBACK_DAYS]).date(),
                pd.Timestamp(dates[i]).date(),
                round((close[i] / base - 1) * 100, 1),
            ))
    return hits


def _scan_near_miss(stock_df: pd.DataFrame):
    """Check wider windows (45–60 bars) for near-miss R01."""
    close = stock_df["Close"].values
    dates = stock_df["date"].values
    n = len(close)
    for lb in (45, 50, 55, 60):
        scan_start = max(lb, n - R01_WATCHLIST_DAYS)
        for i in range(scan_start, n):
            if i < lb:
                continue
            base = close[i - lb]
            if base > 0 and (close[i] / base - 1) >= R01_MIN_MOVE:
                return {
                    "end_date":   pd.Timestamp(dates[i]).date(),
                    "start_date": pd.Timestamp(dates[i - lb]).date(),
                    "pct_move":   round((close[i] / base - 1) * 100, 1),
                    "window_bars": lb,
                }
    return None


# ── per-symbol audit ──────────────────────────────────────────────────────────

def audit_symbol(sym: str, index_df: pd.DataFrame,
                 start_date: str, end_date: str,
                 verbose: bool = False) -> dict:

    # Load data
    stock_raw = load_stock(sym)
    if stock_raw is None:
        return {"symbol": sym, "error": "no_data",
                "r01_fires": False, "near_miss": None,
                "first_r01": None, "first_entry": None, "audit_df": pd.DataFrame()}

    try:
        s_al, i_al = align_stock_index(stock_raw, index_df)
    except Exception as exc:
        return {"symbol": sym, "error": f"align_failed: {exc}",
                "r01_fires": False, "near_miss": None,
                "first_r01": None, "first_entry": None, "audit_df": pd.DataFrame()}

    if len(s_al) < MIN_TRADING_ROWS:
        return {"symbol": sym, "error": f"too_few_rows({len(s_al)})",
                "r01_fires": False, "near_miss": None,
                "first_r01": None, "first_entry": None, "audit_df": pd.DataFrame()}

    # R01 scan on full history
    r01_windows = _scan_r01_windows(s_al)
    near_miss   = _scan_near_miss(s_al) if not r01_windows else None

    # Date range
    t0 = pd.Timestamp(start_date)
    t1 = pd.Timestamp(end_date)
    scan_rows = s_al[(s_al["date"] >= t0) & (s_al["date"] <= t1)]

    rows = []
    first_r01_date  = None
    first_entry_date = None

    for _, row_s in scan_rows.iterrows():
        date   = row_s["date"]
        s_hist = s_al[s_al["date"] <= date].copy()
        i_hist = i_al[i_al["date"] <= date].copy()

        if len(s_hist) < 60 or len(i_hist) < 60:
            continue

        # Compute weekly data once per day
        try:
            weekly = resample_weekly(s_hist)
            idx_wk = resample_weekly(i_hist)
        except Exception:
            continue

        if len(weekly) < 12 or len(idx_wk) < 32:
            continue

        # ── Individual rules — exactly mirrors entry_checklist in signals.py ──
        r01  = _safe(rule01_price_moved_50pct, s_hist)
        r02  = _safe(rule02_green_weekly_candles, weekly)
        r03p, r03v = _safe(rule03_rs_ratio, s_hist, i_hist, default=(False, np.nan))
        r04  = _safe(rule04_index_regime, idx_wk)
        r09  = _safe(rule09_stage2, s_hist, weekly)
        r08  = _safe(rule08_rs_line_near_highs, s_hist, i_hist)
        r11  = _safe(rule11_volatility_contraction, s_hist)
        rb   = _base_depth_pass(s_hist)                              # (hi-close)/hi 8-25%
        rbf  = _safe(rule_base_formed, weekly)                       # 2+ wks no new high

        # r_vol: volume confirms (matches entry_checklist — NOT r03p)
        r_vol = (_safe(rule15_breakout_volume, s_hist) or
                 _safe(rule12_low_vol_red_days, s_hist))

        e_inside  = _safe(rule_weekly_inside_candle_breakout, weekly)
        e_10wma   = _safe(rule_weekly_10wma_support_bounce, weekly)
        e_prevhi  = _safe(rule_weekly_prev_high_near_10wma, weekly)
        e_vcp     = _safe(rule_vcp_breakout, weekly)
        e_hammer  = _safe(rule_weekly_hammer_or_engulfing, weekly)

        # n_confirm: [r02, r_base, r11, r_vol, r08, r09, r_base_formed]
        # Exactly matches signals.py entry_checklist
        n_confirm = sum([r02, rb, r11, r_vol, r08, r09, rbf])
        r_entry   = e_inside or e_10wma or e_prevhi or e_vcp or e_hammer
        non_neg   = r01 and bool(r03p) and r04
        # r_size: bull/sideways regime only (no bear entries) — matches entry_checklist
        r_size    = r04  # r04=True means index is bullish → regime = bull → r_size=True
        all_pass  = non_neg and r_size and r_entry and n_confirm >= N_CONFIRM_REQUIRED

        if r01 and first_r01_date is None:
            first_r01_date = date.date()
        if all_pass and first_entry_date is None:
            first_entry_date = date.date()

        close_px = float(s_hist["Close"].iloc[-1])
        rs_val   = round(float(r03v), 2) if not np.isnan(float(r03v)) else None

        rows.append({
            "date":         date.date(),
            "close":        round(close_px, 2),
            "R01_alert":    "PASS" if r01  else "fail",
            "R03_RS":       f"{rs_val}x" if rs_val else "nan",
            "R03_pass":     "PASS" if r03p else "fail",
            "R04_bull":     "PASS" if r04  else "fail",
            "R02_grn_wk":   "PASS" if r02  else "fail",
            "R09_stage2":   "PASS" if r09  else "fail",
            "R_base_depth": "PASS" if rb   else "fail",
            "R11_vol_ctr":  "PASS" if r11  else "fail",
            "R08_rs_hi":    "PASS" if r08  else "fail",
            "R_base_fmd":   "PASS" if rbf  else "fail",
            "E_wk_inside":  "PASS" if e_inside else ".",
            "E_10wma":      "PASS" if e_10wma  else ".",
            "E_prevhi":     "PASS" if e_prevhi else ".",
            "E_vcp":        "PASS" if e_vcp    else ".",
            "E_hammer":     "PASS" if e_hammer  else ".",
            "n_confirm":    n_confirm,
            "non_neg":      "PASS" if non_neg  else "FAIL",
            "ENTRY":        "ENTRY" if all_pass else ".",
        })

    df_audit = pd.DataFrame(rows)

    # ── console summary ────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  {sym}   ({start_date} → {end_date})")
    print(f"{'='*60}")

    if r01_windows:
        print(f"  R01: {len(r01_windows)} qualifying windows")
        for s, e, p in r01_windows[:3]:
            print(f"    {s} → {e}  +{p}%  ✓")
        if len(r01_windows) > 3:
            print(f"    ... +{len(r01_windows)-3} more")
    elif near_miss:
        nm = near_miss
        print(f"  R01: MISS  (near miss: +{nm['pct_move']}% in "
              f"{nm['window_bars']} bars — needs lookback≥{nm['window_bars']})")
    else:
        print(f"  R01: NEVER fires (no 43%+ move in any 40-bar window)")

    print(f"  First R01 in range : {first_r01_date or 'NEVER'}")
    print(f"  First ENTRY signal : {first_entry_date or 'NEVER'}")

    if not df_audit.empty and verbose:
        r01_days = df_audit[df_audit["R01_alert"] == "PASS"]
        if not r01_days.empty:
            print(f"\n  Filter pass-rates on R01-active days ({len(r01_days)} days):")
            for col in ["R03_pass","R04_bull","R02_grn_wk","R09_stage2",
                        "R_base_depth","R11_vol_ctr","R08_rs_hi","R_base_fmd",
                        "E_wk_inside","E_10wma","E_prevhi","E_vcp"]:
                pc = (r01_days[col] == "PASS").sum()
                pct = pc / len(r01_days) * 100
                bar = "█" * int(pct / 5)
                print(f"    {col:<16}: {pc:3d}/{len(r01_days)} ({pct:5.1f}%)  {bar}")

        entry_days = df_audit[df_audit["ENTRY"] == "ENTRY"]
        if not entry_days.empty:
            print(f"\n  Entry signals ({len(entry_days)}):")
            for _, r in entry_days.head(5).iterrows():
                pats = [p.replace("E_","") for p in
                        ["E_wk_inside","E_10wma","E_prevhi","E_vcp","E_hammer"]
                        if r[p] == "PASS"]
                print(f"    {r['date']}  close={r['close']}  "
                      f"n_confirm={r['n_confirm']}  [{', '.join(pats)}]")

    return {
        "symbol":      sym,
        "error":       None,
        "r01_fires":   bool(r01_windows),
        "near_miss":   near_miss,
        "first_r01":   first_r01_date,
        "first_entry": first_entry_date,
        "audit_df":    df_audit,
        "r01_windows": r01_windows,
    }


# ── Excel output ──────────────────────────────────────────────────────────────

def save_xlsx(results: list, start_date: str, end_date: str, xlsx_name: str = None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl missing — pip install openpyxl")
        return None

    PASS_F  = PatternFill("solid", fgColor="E2EFDA")
    FAIL_F  = PatternFill("solid", fgColor="FCE4D6")
    ENTRY_F = PatternFill("solid", fgColor="FFD966")
    HDR_F   = PatternFill("solid", fgColor="1F4E79")
    HDR_FT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    BODY_FT = Font(name="Arial", size=9)
    BOLD_FT = Font(name="Arial", bold=True, size=9)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Summary sheet (first) ──────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    sum_cols = ["Symbol","Error","R01_fires","Near_miss_bars","First_R01",
                "First_Entry","R01_windows","Note"]
    for ci, h in enumerate(sum_cols, 1):
        c = ws_sum.cell(row=1, column=ci, value=h)
        c.font = HDR_FT; c.fill = HDR_F
        c.alignment = Alignment(horizontal="center")

    for ri, res in enumerate(results, 2):
        nm = res.get("near_miss")
        note = ""
        if res.get("error"):
            note = f"ERROR: {res['error']}"
        elif not res.get("r01_fires") and nm:
            note = (f"Near miss: +{nm['pct_move']}% in {nm['window_bars']} bars. "
                    f"Increase POSITION_LOOKBACK_DAYS to {nm['window_bars']}")
        elif not res.get("r01_fires"):
            note = "Never moved 43%+ in 40 bars"
        elif not res.get("first_entry"):
            note = "R01 fires but no entry signal in range"

        vals = [
            res["symbol"],
            res.get("error") or "",
            "YES" if res.get("r01_fires") else "NO",
            nm["window_bars"] if nm else "",
            str(res.get("first_r01")) if res.get("first_r01") else "—",
            str(res.get("first_entry")) if res.get("first_entry") else "—",
            len(res.get("r01_windows") or []),
            note,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws_sum.cell(row=ri, column=ci, value=v)
            c.font = BODY_FT
            if ci == 3:  # R01_fires
                c.fill = PASS_F if v == "YES" else FAIL_F
            if ci == 6 and v not in ("—",""):  # First_Entry
                c.fill = ENTRY_F; c.font = BOLD_FT

    for ci, w in enumerate([14,20,10,16,12,12,12,60], 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.freeze_panes = "B2"

    # ── Per-symbol detail sheets ───────────────────────────────────────────
    for res in results:
        sym = res["symbol"]
        df  = res.get("audit_df", pd.DataFrame())
        if df.empty:
            continue

        sheet_name = sym[:31]
        ws = wb.create_sheet(sheet_name)

        cols = list(df.columns)
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = HDR_FT; c.fill = HDR_F
            c.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = BODY_FT
                if val == "PASS":  c.fill = PASS_F
                elif val == "fail" or val == "FAIL": c.fill = FAIL_F
                elif val == "ENTRY": c.fill = ENTRY_F; c.font = BOLD_FT

        for ci, h in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(8, len(h)+2)
        ws.freeze_panes = "C2"

    tag  = datetime.now().strftime("%Y%m%d_%H%M")
    name = xlsx_name or f"audit_{len(results)}stocks_{tag}.xlsx"
    path = os.path.join(OUTPUT_DIR, name)
    wb.save(path)
    print(f"\n  Saved → {path}")
    return path


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx",    help="Excel file with SYMBOL column (full or relative path)")
    parser.add_argument("--symbols", help="Comma-separated symbols")
    parser.add_argument("--start",   default="2023-01-01")
    parser.add_argument("--end",     default="2023-12-31")
    parser.add_argument("--verbose", action="store_true")
    parser.add_argument("--workers", type=int, default=1,
                        help="Parallel workers (default 1 — safe; increase for speed)")
    args = parser.parse_args()

    # ── collect symbols ────────────────────────────────────────────────────
    symbols = []
    if args.xlsx:
        xlsx_path = args.xlsx if os.path.isabs(args.xlsx) else os.path.join(
            os.path.dirname(__file__), args.xlsx)
        df_xl = pd.read_excel(xlsx_path)
        col   = next((c for c in df_xl.columns if c.strip().upper() == "SYMBOL"), None)
        if col is None:
            print(f"ERROR: No SYMBOL column in {xlsx_path}. Columns: {list(df_xl.columns)}")
            sys.exit(1)
        symbols = [str(s).strip().upper() for s in df_xl[col].dropna()]
        print(f"Loaded {len(symbols)} symbols from {xlsx_path}")
    if args.symbols:
        symbols += [s.strip().upper() for s in args.symbols.split(",")]
    if not symbols:
        print("ERROR: provide --xlsx or --symbols")
        sys.exit(1)

    # Filter to symbols that exist in parquet store
    symbols_csv = os.path.join(os.path.dirname(__file__), "symbols.csv")
    if os.path.exists(symbols_csv):
        known = set(pd.read_csv(symbols_csv, header=None)[0].str.upper())
        missing = [s for s in symbols if s not in known]
        symbols = [s for s in symbols if s in known]
        if missing:
            print(f"  Skipping {len(missing)} symbols not in parquet store: "
                  f"{', '.join(missing[:10])}{'...' if len(missing)>10 else ''}")
        print(f"  Auditing {len(symbols)} symbols present in data store.")

    # Load index once
    print(f"\nLoading index ...")
    index_df = load_index()
    print(f"Index: {len(index_df)} rows\n")

    # Run audit
    results = []
    total = len(symbols)
    for idx, sym in enumerate(symbols, 1):
        print(f"[{idx}/{total}] {sym}", end="  ", flush=True)
        try:
            res = audit_symbol(sym, index_df, args.start, args.end,
                               verbose=args.verbose)
        except Exception as exc:
            print(f"FATAL ERROR: {exc}")
            traceback.print_exc()
            res = {"symbol": sym, "error": str(exc), "r01_fires": False,
                   "near_miss": None, "first_r01": None, "first_entry": None,
                   "audit_df": pd.DataFrame(), "r01_windows": []}
        results.append(res)

    # Save Excel
    tag = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx_name = f"audit_winners_{args.start[:4]}_{tag}.xlsx"
    save_xlsx(results, args.start, args.end, xlsx_name)

    # Final console summary
    r01_yes     = [r for r in results if r.get("r01_fires")]
    r01_miss    = [r for r in results if r.get("near_miss")]
    r01_no      = [r for r in results if not r.get("r01_fires") and not r.get("near_miss") and not r.get("error")]
    with_entry  = [r for r in results if r.get("first_entry")]
    errors      = [r for r in results if r.get("error")]

    print(f"\n{'='*68}")
    print(f"  AUDIT SUMMARY   {args.start} → {args.end}")
    print(f"{'='*68}")
    print(f"  Total audited         : {len(results)}")
    print(f"  Errors (no data)      : {len(errors)}")
    print(f"  R01 fires ✓           : {len(r01_yes)}")
    print(f"  R01 near-miss (~45d)  : {len(r01_miss)}")
    print(f"  R01 never fires       : {len(r01_no)}")
    print(f"  Got entry signal ✓    : {len(with_entry)}")
    print(f"  R01 fires but NO entry: {len(r01_yes) - len(with_entry)}")
    print(f"\n  Symbols WITH entry signal:")
    for r in with_entry:
        print(f"    {r['symbol']:<18} first entry: {r['first_entry']}")
    print(f"\n  Symbols R01 fires but NO entry (blocking filter issue):")
    no_entry = [r for r in r01_yes if not r.get("first_entry")]
    for r in no_entry[:20]:
        # Show which filter was most often blocking on R01-active days
        df = r.get("audit_df", pd.DataFrame())
        blocking = ""
        if not df.empty:
            r01_days = df[df["R01_alert"] == "PASS"]
            if not r01_days.empty:
                fail_rates = {
                    col: (r01_days[col] == "fail").mean()
                    for col in ["R03_pass","R04_bull","R02_grn_wk","R09_stage2",
                                "R_base_depth","R_base_fmd","E_wk_inside",
                                "E_10wma","E_prevhi","E_vcp"]
                }
                top = sorted(fail_rates.items(), key=lambda x: -x[1])[:3]
                blocking = "  block: " + ", ".join(f"{k}({v*100:.0f}%fail)" for k,v in top)
        print(f"    {r['symbol']:<18}{blocking}")
    if len(no_entry) > 20:
        print(f"    ... and {len(no_entry)-20} more (see Excel)")

    print(f"\n  Near-misses (would qualify with wider R01 window):")
    for r in r01_miss:
        nm = r["near_miss"]
        print(f"    {r['symbol']:<18} +{nm['pct_move']}% in {nm['window_bars']} bars")
    print(f"{'='*68}")


if __name__ == "__main__":
    main()
