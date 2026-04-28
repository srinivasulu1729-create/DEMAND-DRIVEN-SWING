"""
tradelog_validator.py — Demand-Driven Swing & Position System

Cross-checks the manual TradeLog.xlsx against system-generated signals.
Output: tradelog_validation.xlsx
"""

import logging
import os
from datetime import timedelta

import numpy as np
import pandas as pd

from config import TRADELOG_XLSX, OUTPUT_DIR
from data_loader import load_index, load_stock, align_stock_index
from indicators import resample_weekly, rs_ratio, vol_ratio, index_regime
from signals import (
    entry_checklist, rule03_rs_ratio,
    sell_on_strength, sell_when_extended, sell_on_weakness,
)

logger = logging.getLogger(__name__)
TOLERANCE_DAYS = 3


def _parse_tradelog(path: str = TRADELOG_XLSX) -> pd.DataFrame:
    """Load and clean the TradeLog.xlsx."""
    df = pd.read_excel(path, sheet_name="Sheet1")
    df.columns = df.columns.str.strip()

    df = df.rename(columns={
        "Entry date":        "entry_date",
        "Exit date":         "exit_date",
        "Symbol":            "symbol",
        "Entry Price":       "entry_price",
        "Exit price":        "exit_price",
        "Stop loss":         "stop_loss",
        "Setup":             "setup",
        "Exit reason":       "exit_reason",
        "Alert Range start": "alert_start",
        "Alert Range end":   "alert_end",
    })

    def safe_date(x):
        if pd.isna(x):
            return pd.NaT
        try:
            return pd.to_datetime(x)
        except Exception:
            return pd.NaT

    df["entry_date"] = df["entry_date"].apply(safe_date)
    df["exit_date"]  = df["exit_date"].apply(safe_date)
    df["manual_pnl_pct"] = (
        (df["exit_price"] - df["entry_price"]) / df["entry_price"] * 100
    ).round(2)
    return df


def _nearest_date(target, available, tol=TOLERANCE_DAYS):
    mask = (available >= target - timedelta(days=tol)) & \
           (available <= target + timedelta(days=tol))
    cands = available[mask]
    if cands.empty:
        return None
    return cands.iloc[(cands - target).abs().argsort().iloc[0]]


def validate_trade(row, index_df):
    """Validate a single trade row against system signals."""
    sym = row["symbol"]
    entry_date = row["entry_date"]
    exit_date  = row["exit_date"]

    result = {
        "symbol":             sym,
        "setup":              str(row.get("setup", "")),
        "manual_entry_date":  str(entry_date.date()) if not pd.isna(entry_date) else "N/A",
        "manual_exit_date":   str(exit_date.date())  if not pd.isna(exit_date)  else "N/A",
        "manual_entry_price": row["entry_price"],
        "manual_exit_price":  row["exit_price"],
        "manual_pnl_pct":     row["manual_pnl_pct"],
        "sys_entry_signal":   "No",
        "sys_exit_signal":    "No",
        "sys_pnl_pct":        None,
        "checklist_all_pass": None,
        "rs_at_entry":        None,
        "vol_ratio_at_entry": None,
        "regime_at_entry":    None,
        "match_status":       "UNMATCHED",
        "notes":              "",
    }

    if pd.isna(entry_date):
        result["notes"] = "Missing entry date"
        return result

    try:
        stock_df = load_stock(sym)
    except Exception as e:
        result["notes"] = f"Load error: {e}"
        return result

    if stock_df is None:
        result["notes"] = "Symbol not found"
        return result

    stock_df, idx_aligned = align_stock_index(stock_df, index_df)
    avail = stock_df["date"]

    # ── Entry check ───────────────────────────────────────────────────────
    sys_entry_date = _nearest_date(entry_date, avail)
    if sys_entry_date is None:
        result["notes"] = "No data near entry date"
        return result

    s_hist   = stock_df[stock_df["date"] <= sys_entry_date]
    i_hist   = idx_aligned[idx_aligned["date"] <= sys_entry_date]
    idx_hist = index_df[index_df["date"] <= sys_entry_date]

    if len(s_hist) < 60:
        result["notes"] = "Insufficient history at entry"
        return result

    weekly   = resample_weekly(s_hist)
    idx_wkly = resample_weekly(idx_hist)

    try:
        checklist = entry_checklist(s_hist, i_hist, weekly, idx_wkly, "position")
        _, rs_val = rule03_rs_ratio(s_hist, i_hist)
        vr        = float(vol_ratio(s_hist).iloc[-1])
        regime    = index_regime(idx_wkly)

        result["checklist_all_pass"]  = checklist["all_pass"]
        result["rs_at_entry"]         = round(rs_val, 2) if not np.isnan(rs_val) else "N/A"
        result["vol_ratio_at_entry"]  = round(vr, 2)
        result["regime_at_entry"]     = regime

        # Individual checklist items
        for k in ["r01_price_50pct","r02_green_weekly","r03_rs_ratio",
                  "r04_index_regime","r_base_depth_8_25pct","r11_vol_contracting",
                  "r_entry_valid","r_volume_confirms","r_position_sized","r08_rs_near_highs"]:
            result[f"cl_{k}"] = checklist.get(k, None)

        if checklist["non_negotiable_pass"]:
            result["sys_entry_signal"] = "YES-AllPass" if checklist["all_pass"] else "YES-NonNeg"
        else:
            result["sys_entry_signal"] = "PARTIAL"
    except Exception as e:
        result["notes"] += f" Entry err: {e}"

    # ── Exit check ────────────────────────────────────────────────────────
    if not pd.isna(exit_date):
        sys_exit_date = _nearest_date(exit_date, avail)
        if sys_exit_date is not None:
            s_exit = stock_df[stock_df["date"] <= sys_exit_date]
            i_exit = idx_aligned[idx_aligned["date"] <= sys_exit_date]
            if len(s_exit) >= 20:
                try:
                    entry_px   = float(row["entry_price"])
                    exit_close = float(s_exit["Close"].iloc[-1])
                    sys_pnl    = (exit_close - entry_px) / entry_px * 100
                    sos = sell_on_strength(s_exit, entry_px)
                    swe = sell_when_extended(s_exit)
                    sow = sell_on_weakness(s_exit, i_exit)
                    exit_ok = sos["sell_strength"] or swe["sell_extended"] or sow["sell_weakness"]
                    result["sys_exit_signal"] = "YES" if exit_ok else "No"
                    result["sys_pnl_pct"]     = round(sys_pnl, 2)
                except Exception as e:
                    result["notes"] += f" Exit err: {e}"

    # ── Match status ──────────────────────────────────────────────────────
    e_ok = "YES" in result["sys_entry_signal"]
    x_ok = result["sys_exit_signal"] == "YES"
    if e_ok and x_ok:
        result["match_status"] = "FULL MATCH"
    elif e_ok:
        result["match_status"] = "ENTRY MATCH"
    elif x_ok:
        result["match_status"] = "EXIT MATCH"
    else:
        result["match_status"] = "NO MATCH"

    return result


def run_validation(output_dir: str = OUTPUT_DIR) -> pd.DataFrame:
    """Validate all trades in TradeLog.xlsx. Saves tradelog_validation.xlsx."""
    logger.info("Loading trade log: %s", TRADELOG_XLSX)
    trade_log = _parse_tradelog()
    logger.info("Validating %d trades ...", len(trade_log))

    index_df = load_index()
    rows = []
    for i, row in trade_log.iterrows():
        logger.info("[%d/%d] %s ...", i + 1, len(trade_log), row["symbol"])
        rows.append(validate_trade(row, index_df))

    val_df = pd.DataFrame(rows)

    full_match  = (val_df["match_status"] == "FULL MATCH").sum()
    entry_match = (val_df["match_status"] == "ENTRY MATCH").sum()
    exit_match  = (val_df["match_status"] == "EXIT MATCH").sum()
    no_match    = (val_df["match_status"] == "NO MATCH").sum()

    print(f"\n{'='*60}")
    print("  TRADE LOG VALIDATION SUMMARY")
    print(f"{'='*60}")
    print(f"  Total trades      : {len(val_df)}")
    print(f"  Full match        : {full_match}")
    print(f"  Entry match only  : {entry_match}")
    print(f"  Exit match only   : {exit_match}")
    print(f"  No match          : {no_match}")

    matched = val_df[val_df["sys_pnl_pct"].notna()].copy()
    if not matched.empty:
        matched["pnl_diff"] = matched["sys_pnl_pct"] - matched["manual_pnl_pct"]
        print(f"  Avg manual P&L    : {matched['manual_pnl_pct'].mean():.2f}%")
        print(f"  Avg system P&L    : {matched['sys_pnl_pct'].mean():.2f}%")
        print(f"  Avg P&L diff      : {matched['pnl_diff'].mean():.2f}%")
    print(f"{'='*60}\n")

    out_path = os.path.join(output_dir, "tradelog_validation.xlsx")
    try:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            val_df.to_excel(writer, sheet_name="Validation", index=False)
            ws = writer.sheets["Validation"]
            from openpyxl.styles import PatternFill
            green_fill  = PatternFill("solid", fgColor="C6EFCE")
            yellow_fill = PatternFill("solid", fgColor="FFEB9C")
            red_fill    = PatternFill("solid", fgColor="FFC7CE")
            headers = [c.value for c in ws[1]]
            ms_col = None
            try:
                ms_col = headers.index("match_status") + 1
            except ValueError:
                pass
            if ms_col:
                for row in ws.iter_rows(min_row=2, min_col=ms_col, max_col=ms_col):
                    for cell in row:
                        if cell.value == "FULL MATCH":
                            cell.fill = green_fill
                        elif cell.value in ("ENTRY MATCH", "EXIT MATCH"):
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        print(f"  Saved: {out_path}")
    except Exception as e:
        logger.error("xlsx save failed: %s", e)
        csv_path = out_path.replace(".xlsx", ".csv")
        val_df.to_csv(csv_path, index=False)
        print(f"  Saved (CSV): {csv_path}")

    return val_df
