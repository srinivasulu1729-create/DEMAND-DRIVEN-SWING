"""
audit_report.py — Demand-Driven Swing & Position System

Generates a comprehensive multi-sheet Excel audit report by cross-checking
every trade in TradeLog.xlsx against system signals.

Sheets produced:
  1. Summary          — headline stats, win rates, P&L comparison
  2. Trade_Detail     — per-trade: manual vs system, entry/exit signals, P&L diff
  3. Rule_Breakdown   — all 10 checklist rules per trade (pass/fail heatmap)
  4. Symbol_Summary   — per-symbol aggregation (trades, win rate, avg P&L)
  5. PnL_Comparison   — manual P&L vs system P&L scatter-ready table
  6. Missed_Signals   — trades where system WOULD NOT have fired (with reasons)
  7. Extra_Signals    — dates where system fires but no manual trade exists

Run:
    python audit_report.py
    python main.py --validate  (calls run_audit internally)
"""

import logging
import os
from datetime import timedelta, datetime

import numpy as np
import pandas as pd

from config import TRADELOG_XLSX, OUTPUT_DIR
from data_loader import load_index, load_stock, align_stock_index
from indicators import resample_weekly, rs_ratio, vol_ratio, atr, index_regime, rsi
from signals import (
    entry_checklist, rule03_rs_ratio,
    rule01_price_moved_50pct, rule02_green_weekly_candles,
    rule04_index_regime, rule05_low_vol_pullbacks,
    rule06_high_vol_expansions, rule07_atr_contraction,
    rule08_rs_line_near_highs, rule09_stage2,
    rule10_base_bottom_entry, rule11_volatility_contraction,
    rule14_new_high_breakout, rule15_breakout_volume,
    sell_on_strength, sell_when_extended, sell_on_weakness,
)

logger = logging.getLogger(__name__)
TOLERANCE_DAYS = 3

CHECKLIST_RULES = {
    "R01  Price +50% in 40d":       "r01_price_50pct",
    "R02  ≥6 Green Weekly Candles": "r02_green_weekly",
    "R03  RS ≥ 3× Index":           "r03_rs_ratio",
    "R04  Index Above EMAs":        "r04_index_regime",
    "R    Base Depth 8–25%":        "r_base_depth_8_25pct",
    "R11  Volatility Contracting":  "r11_vol_contracting",
    "R    Entry Valid (BB/BO)":      "r_entry_valid",
    "R    Volume Confirms":         "r_volume_confirms",
    "R    Position Sized OK":       "r_position_sized",
    "R08  RS Line Near Highs":      "r08_rs_near_highs",
}

EXIT_RULES = {
    "SOS R25 +20% Target":          "rule25_target_hit",
    "SOS R26 Wide Candle+Vol":      "rule26_wide_range_vol",
    "SOS R27 40% Above 10W EMA":    "rule27_extended_10W",
    "SOS R28 Ignite Bar":           "rule28_ignite_bar",
    "SWE R29 Accel Days":           "rule29_accel_days",
    "SWE R30 RSI > 85":             "rule30_rsi_overbought",
    "SOW R31 Below 20EMA+Vol":      "rule31_close_below_20ema",
    "SOW R32 Lower Low":            "rule32_lower_low",
    "SOW R33 RS Breakdown":         "rule33_rs_breakdown",
    "SOW R34 Heavy Red Vol×2":      "rule34_heavy_red_vol",
}


# ═══════════════════════════════════════════════════════════════════════════
# PARSE TRADE LOG
# ═══════════════════════════════════════════════════════════════════════════

def _parse_tradelog(path: str = TRADELOG_XLSX) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Sheet1")
    df.columns = df.columns.str.strip()
    df = df.rename(columns={
        "Entry date":        "entry_date",
        "Exit date":         "exit_date",
        "Symbol":            "symbol",
        "Entry Price":       "entry_price",
        "Exit price":        "exit_price",
        "Stop loss":         "stop_loss",
        "Setup":             "setup",
        "Exit reason":       "exit_reason",
        "Alert Range start": "alert_start",
        "Alert Range end":   "alert_end",
        "Alert Condition":   "alert_condition",
    })

    def safe_date(x):
        if pd.isna(x):
            return pd.NaT
        try:
            return pd.to_datetime(x)
        except Exception:
            return pd.NaT

    df["entry_date"] = df["entry_date"].apply(safe_date)
    df["exit_date"]  = df["exit_date"].apply(safe_date)
    df["manual_pnl_pct"] = (
        (df["exit_price"] - df["entry_price"]) / df["entry_price"] * 100
    ).round(2)
    df["manual_rr"] = (df["exit_price"] - df["entry_price"]) / (
        df["entry_price"] - df["stop_loss"]
    ).replace(0, np.nan)
    df["trade_id"] = range(1, len(df) + 1)
    return df


def _nearest_date(target, available, tol=TOLERANCE_DAYS):
    if pd.isna(target):
        return None
    mask  = (available >= target - timedelta(days=tol)) & \
            (available <= target + timedelta(days=tol))
    cands = available[mask]
    if cands.empty:
        return None
    return cands.iloc[(cands - target).abs().argsort().iloc[0]]


# ═══════════════════════════════════════════════════════════════════════════
# PER-TRADE AUDIT
# ═══════════════════════════════════════════════════════════════════════════

def _audit_trade(row: pd.Series, index_df: pd.DataFrame) -> dict:
    sym        = row["symbol"]
    entry_date = row["entry_date"]
    exit_date  = row["exit_date"]

    base = {
        "trade_id":           row["trade_id"],
        "symbol":             sym,
        "setup":              str(row.get("setup", "")),
        "manual_entry_date":  str(entry_date.date()) if not pd.isna(entry_date) else "N/A",
        "manual_exit_date":   str(exit_date.date())  if not pd.isna(exit_date)  else "N/A",
        "manual_entry_price": row["entry_price"],
        "manual_stop_loss":   row["stop_loss"],
        "manual_exit_price":  row["exit_price"],
        "manual_pnl_pct":     row["manual_pnl_pct"],
        "manual_rr":          round(row["manual_rr"], 2) if not pd.isna(row["manual_rr"]) else None,
        "exit_reason":        str(row.get("exit_reason", "")),
        # System fields (filled in below)
        "sys_entry_found":    False,
        "sys_entry_date":     "N/A",
        "sys_all_pass":       None,
        "sys_non_neg_pass":   None,
        "sys_checklist_score": None,
        "sys_rs_at_entry":    None,
        "sys_vol_ratio_entry": None,
        "sys_atr_at_entry":   None,
        "sys_rsi_at_entry":   None,
        "sys_regime":         None,
        "sys_exit_found":     False,
        "sys_exit_date":      "N/A",
        "sys_exit_price":     None,
        "sys_pnl_pct":        None,
        "pnl_diff_pct":       None,
        "match_status":       "NO MATCH",
        "notes":              "",
    }
    # Add per-rule columns (checklist)
    for label in CHECKLIST_RULES:
        base[label] = None
    # Add per-rule columns (exit)
    for label in EXIT_RULES:
        base[label] = None

    if pd.isna(entry_date):
        base["notes"] = "Missing entry date"
        return base

    # Load stock
    try:
        stock_df = load_stock(sym)
    except Exception as e:
        base["notes"] = f"Load error: {e}"
        return base
    if stock_df is None:
        base["notes"] = "Symbol not in data"
        return base

    stock_df, idx_aligned = align_stock_index(stock_df, index_df)
    avail = stock_df["date"]

    # ── ENTRY AUDIT ───────────────────────────────────────────────────────
    sys_edate = _nearest_date(entry_date, avail)
    if sys_edate is None:
        base["notes"] = "No trading data near entry date"
        return base

    s_hist   = stock_df[stock_df["date"] <= sys_edate]
    i_hist   = idx_aligned[idx_aligned["date"] <= sys_edate]
    idx_hist = index_df[index_df["date"] <= sys_edate]

    if len(s_hist) < 60:
        base["notes"] = "< 60 rows at entry date"
        return base

    weekly   = resample_weekly(s_hist)
    idx_wkly = resample_weekly(idx_hist)

    try:
        checklist = entry_checklist(s_hist, i_hist, weekly, idx_wkly, "position")
        _, rs_val = rule03_rs_ratio(s_hist, i_hist)
        vr        = float(vol_ratio(s_hist).iloc[-1])
        atr_val   = float(atr(s_hist).iloc[-1])
        rsi_val   = float(rsi(s_hist["Close"]).iloc[-1])
        regime    = index_regime(idx_wkly)

        base["sys_entry_found"]     = checklist["non_negotiable_pass"]
        base["sys_entry_date"]      = str(sys_edate.date())
        base["sys_all_pass"]        = checklist["all_pass"]
        base["sys_non_neg_pass"]    = checklist["non_negotiable_pass"]
        base["sys_checklist_score"] = sum(
            1 for k in CHECKLIST_RULES.values() if checklist.get(k, False)
        )
        base["sys_rs_at_entry"]     = round(rs_val, 2) if not np.isnan(rs_val) else None
        base["sys_vol_ratio_entry"] = round(vr, 2)
        base["sys_atr_at_entry"]    = round(atr_val, 2)
        base["sys_rsi_at_entry"]    = round(rsi_val, 2)
        base["sys_regime"]          = regime

        # Fill per-rule checklist columns
        for label, key in CHECKLIST_RULES.items():
            base[label] = checklist.get(key, False)

    except Exception as e:
        base["notes"] += f" Entry err: {e}"

    # ── EXIT AUDIT ────────────────────────────────────────────────────────
    if not pd.isna(exit_date):
        sys_xdate = _nearest_date(exit_date, avail)
        if sys_xdate is not None:
            s_exit = stock_df[stock_df["date"] <= sys_xdate]
            i_exit = idx_aligned[idx_aligned["date"] <= sys_xdate]
            if len(s_exit) >= 20:
                try:
                    ep          = float(row["entry_price"])
                    exit_close  = float(s_exit["Close"].iloc[-1])
                    sys_pnl     = (exit_close - ep) / ep * 100

                    # Weekly data for sell-on-strength
                    wkly_exit   = resample_weekly(s_exit)
                    sos = sell_on_strength(s_exit, ep, wkly_exit if len(wkly_exit) > 5 else None)
                    swe = sell_when_extended(s_exit)
                    sow = sell_on_weakness(s_exit, i_exit)

                    exit_triggered = (sos["sell_strength"] or swe["sell_extended"]
                                      or sow["sell_weakness"])

                    base["sys_exit_found"]  = exit_triggered
                    base["sys_exit_date"]   = str(sys_xdate.date())
                    base["sys_exit_price"]  = round(exit_close, 2)
                    base["sys_pnl_pct"]     = round(sys_pnl, 2)
                    base["pnl_diff_pct"]    = round(sys_pnl - row["manual_pnl_pct"], 2)

                    # Per-exit-rule columns
                    all_exit = {}
                    all_exit.update(sos)
                    all_exit.update(swe)
                    all_exit.update(sow)
                    for label, key in EXIT_RULES.items():
                        base[label] = all_exit.get(key, False)

                except Exception as e:
                    base["notes"] += f" Exit err: {e}"

    # ── Match status ──────────────────────────────────────────────────────
    e_ok = base["sys_entry_found"]
    x_ok = base["sys_exit_found"]
    if e_ok and x_ok:
        base["match_status"] = "FULL MATCH"
    elif e_ok:
        base["match_status"] = "ENTRY MATCH"
    elif x_ok:
        base["match_status"] = "EXIT MATCH"
    else:
        base["match_status"] = "NO MATCH"

    return base


# ═══════════════════════════════════════════════════════════════════════════
# REPORT BUILDER
# ═══════════════════════════════════════════════════════════════════════════

def run_audit(output_dir: str = OUTPUT_DIR) -> pd.DataFrame:
    """
    Build the full multi-sheet audit report.
    Returns the trade-detail DataFrame.
    """
    logger.info("Loading trade log …")
    trade_log = _parse_tradelog()
    logger.info("Loading index …")
    index_df  = load_index()

    # ── Per-trade audit ───────────────────────────────────────────────────
    records = []
    for i, row in trade_log.iterrows():
        logger.info("[%d/%d] Auditing %s  %s …",
                    i + 1, len(trade_log), row["symbol"],
                    str(row["entry_date"].date()) if not pd.isna(row["entry_date"]) else "")
        records.append(_audit_trade(row, index_df))

    detail_df = pd.DataFrame(records)

    # ── Sheet 1 — Summary ─────────────────────────────────────────────────
    total    = len(detail_df)
    fm       = (detail_df["match_status"] == "FULL MATCH").sum()
    em       = (detail_df["match_status"] == "ENTRY MATCH").sum()
    xm       = (detail_df["match_status"] == "EXIT MATCH").sum()
    nm       = (detail_df["match_status"] == "NO MATCH").sum()
    m_wins   = (detail_df["manual_pnl_pct"] > 0).sum()
    m_wr     = m_wins / total * 100
    m_avg    = detail_df["manual_pnl_pct"].mean()

    pnl_df   = detail_df[detail_df["sys_pnl_pct"].notna()].copy()
    s_avg    = pnl_df["sys_pnl_pct"].mean() if not pnl_df.empty else None
    s_wins   = (pnl_df["sys_pnl_pct"] > 0).sum() if not pnl_df.empty else 0
    s_wr     = s_wins / len(pnl_df) * 100 if not pnl_df.empty else None
    avg_diff = pnl_df["pnl_diff_pct"].mean() if not pnl_df.empty else None

    summary_data = {
        "Metric": [
            "Total Trades (Manual)",
            "Full Match (Entry + Exit)",
            "Entry Match Only",
            "Exit Match Only",
            "No Match",
            "Match Rate %",
            "─── Manual Log Stats ───",
            "Manual Win Rate %",
            "Manual Avg P&L %",
            "Manual Best Trade",
            "Manual Worst Trade",
            "Manual Best Symbol",
            "─── System Comparison ───",
            "System Win Rate % (matched)",
            "System Avg P&L % (matched)",
            "Avg P&L Diff % (sys − manual)",
            "Trades where System > Manual",
            "Trades where Manual > System",
        ],
        "Value": [
            total,
            f"{fm}  ({fm/total*100:.1f}%)",
            f"{em}  ({em/total*100:.1f}%)",
            f"{xm}  ({xm/total*100:.1f}%)",
            f"{nm}  ({nm/total*100:.1f}%)",
            f"{(fm+em)/total*100:.1f}%",
            "",
            f"{m_wr:.1f}%",
            f"{m_avg:.2f}%",
            f"{detail_df['manual_pnl_pct'].max():.2f}%",
            f"{detail_df['manual_pnl_pct'].min():.2f}%",
            detail_df.loc[detail_df["manual_pnl_pct"].idxmax(), "symbol"],
            "",
            f"{s_wr:.1f}%" if s_wr else "N/A",
            f"{s_avg:.2f}%" if s_avg else "N/A",
            f"{avg_diff:.2f}%" if avg_diff else "N/A",
            str((pnl_df["pnl_diff_pct"] > 0).sum()) if not pnl_df.empty else "N/A",
            str((pnl_df["pnl_diff_pct"] < 0).sum()) if not pnl_df.empty else "N/A",
        ],
    }
    summary_df = pd.DataFrame(summary_data)

    # ── Sheet 4 — Symbol Summary ──────────────────────────────────────────
    sym_grp = detail_df.groupby("symbol").agg(
        Trades=("trade_id", "count"),
        ManualWins=("manual_pnl_pct", lambda x: (x > 0).sum()),
        ManualAvgPnL=("manual_pnl_pct", "mean"),
        ManualMaxPnL=("manual_pnl_pct", "max"),
        ManualMinPnL=("manual_pnl_pct", "min"),
        SysAvgPnL=("sys_pnl_pct", "mean"),
        AvgPnLDiff=("pnl_diff_pct", "mean"),
        FullMatch=("match_status", lambda x: (x == "FULL MATCH").sum()),
        AvgChecklistScore=("sys_checklist_score", "mean"),
        AvgRS=("sys_rs_at_entry", "mean"),
    ).reset_index()
    sym_grp["ManualWinRate%"] = (sym_grp["ManualWins"] / sym_grp["Trades"] * 100).round(1)
    sym_grp["ManualAvgPnL"]   = sym_grp["ManualAvgPnL"].round(2)
    sym_grp["MatchRate%"]     = (sym_grp["FullMatch"] / sym_grp["Trades"] * 100).round(1)
    sym_grp = sym_grp.sort_values("ManualAvgPnL", ascending=False)

    # ── Sheet 5 — P&L comparison table ────────────────────────────────────
    pnl_compare = detail_df[[
        "trade_id", "symbol", "setup",
        "manual_entry_date", "manual_entry_price",
        "manual_exit_price", "manual_pnl_pct",
        "sys_pnl_pct", "pnl_diff_pct", "match_status",
    ]].copy()

    # ── Sheet 6 — Missed signals (NO MATCH) ───────────────────────────────
    missed = detail_df[detail_df["match_status"] == "NO MATCH"][[
        "trade_id", "symbol", "setup", "manual_entry_date",
        "manual_entry_price", "manual_pnl_pct",
        "sys_non_neg_pass", "sys_checklist_score",
        "sys_rs_at_entry", "sys_regime", "notes",
    ] + list(CHECKLIST_RULES.keys())].copy()

    # ── Sheet 3 — Rule breakdown ──────────────────────────────────────────
    rule_cols = ["trade_id", "symbol", "manual_entry_date",
                 "sys_checklist_score", "sys_all_pass", "match_status"]
    rule_cols += list(CHECKLIST_RULES.keys())
    rule_cols += list(EXIT_RULES.keys())
    rule_df = detail_df[[c for c in rule_cols if c in detail_df.columns]].copy()

    # ── PRINT console summary ─────────────────────────────────────────────
    print(f"\n{'═'*65}")
    print("  AUDIT REPORT SUMMARY")
    print(f"{'═'*65}")
    for _, r in summary_df.iterrows():
        if str(r["Value"]).startswith("─"):
            print(f"  {r['Metric']}")
        else:
            print(f"  {r['Metric']:<40} {r['Value']}")
    print(f"{'═'*65}\n")

    # ── WRITE EXCEL ───────────────────────────────────────────────────────
    stamp    = datetime.today().strftime("%Y%m%d_%H%M")
    out_path = os.path.join(output_dir, f"audit_report_{stamp}.xlsx")

    try:
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter

        green_fill  = PatternFill("solid", fgColor="C6EFCE")
        red_fill    = PatternFill("solid", fgColor="FFC7CE")
        yellow_fill = PatternFill("solid", fgColor="FFEB9C")
        blue_fill   = PatternFill("solid", fgColor="DDEBF7")
        hdr_fill    = PatternFill("solid", fgColor="1F4E79")
        hdr_font    = Font(color="FFFFFF", bold=True)
        bold_font   = Font(bold=True)

        def style_header(ws):
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[1].height = 30

        def autofit(ws, max_width=45):
            for col in ws.columns:
                mx = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(mx + 2, max_width)

        def colour_match_col(ws, col_idx):
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value == "FULL MATCH":
                        cell.fill = green_fill
                    elif cell.value in ("ENTRY MATCH", "EXIT MATCH"):
                        cell.fill = yellow_fill
                    elif cell.value == "NO MATCH":
                        cell.fill = red_fill

        def colour_bool_cols(ws, col_indices):
            for ci in col_indices:
                for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                    for cell in row:
                        if cell.value is True:
                            cell.fill = green_fill
                        elif cell.value is False:
                            cell.fill = red_fill

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

            # Sheet 1: Summary
            summary_df.to_excel(writer, sheet_name="1_Summary", index=False)
            ws = writer.sheets["1_Summary"]
            style_header(ws)
            autofit(ws)
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value or "").startswith("─"):
                    for cell in row:
                        cell.font   = bold_font
                        cell.fill   = blue_fill

            # Sheet 2: Trade Detail
            detail_cols = [
                "trade_id", "symbol", "setup",
                "manual_entry_date", "manual_entry_price",
                "manual_stop_loss", "manual_exit_date",
                "manual_exit_price", "manual_pnl_pct", "manual_rr",
                "exit_reason",
                "sys_entry_found", "sys_entry_date",
                "sys_all_pass", "sys_checklist_score",
                "sys_rs_at_entry", "sys_vol_ratio_entry",
                "sys_atr_at_entry", "sys_rsi_at_entry", "sys_regime",
                "sys_exit_found", "sys_exit_date",
                "sys_exit_price", "sys_pnl_pct",
                "pnl_diff_pct", "match_status", "notes",
            ]
            show_cols = [c for c in detail_cols if c in detail_df.columns]
            detail_df[show_cols].to_excel(writer, sheet_name="2_Trade_Detail", index=False)
            ws2 = writer.sheets["2_Trade_Detail"]
            style_header(ws2)
            hdr2 = [c.value for c in ws2[1]]
            ms_col2 = hdr2.index("match_status") + 1 if "match_status" in hdr2 else None
            if ms_col2:
                colour_match_col(ws2, ms_col2)
            bool_cols2 = [hdr2.index(c) + 1 for c in ["sys_entry_found","sys_exit_found","sys_all_pass"]
                          if c in hdr2]
            colour_bool_cols(ws2, bool_cols2)
            # Colour pnl_diff
            if "pnl_diff_pct" in hdr2:
                diff_col = hdr2.index("pnl_diff_pct") + 1
                for row in ws2.iter_rows(min_row=2, min_col=diff_col, max_col=diff_col):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.fill = green_fill if cell.value >= 0 else red_fill
            autofit(ws2)

            # Sheet 3: Rule Breakdown (heatmap)
            rule_df.to_excel(writer, sheet_name="3_Rule_Breakdown", index=False)
            ws3 = writer.sheets["3_Rule_Breakdown"]
            style_header(ws3)
            hdr3 = [c.value for c in ws3[1]]
            cl_bool_idx = [i + 1 for i, h in enumerate(hdr3)
                           if h in list(CHECKLIST_RULES.keys()) + list(EXIT_RULES.keys())]
            colour_bool_cols(ws3, cl_bool_idx)
            ms_col3 = hdr3.index("match_status") + 1 if "match_status" in hdr3 else None
            if ms_col3:
                colour_match_col(ws3, ms_col3)
            autofit(ws3, max_width=35)

            # Sheet 4: Symbol Summary
            sym_grp.to_excel(writer, sheet_name="4_Symbol_Summary", index=False)
            ws4 = writer.sheets["4_Symbol_Summary"]
            style_header(ws4)
            autofit(ws4)
            # Colour ManualWinRate%
            hdr4 = [c.value for c in ws4[1]]
            wr_col = hdr4.index("ManualWinRate%") + 1 if "ManualWinRate%" in hdr4 else None
            if wr_col:
                for row in ws4.iter_rows(min_row=2, min_col=wr_col, max_col=wr_col):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.fill = green_fill if cell.value >= 60 else yellow_fill

            # Sheet 5: P&L Comparison
            pnl_compare.to_excel(writer, sheet_name="5_PnL_Comparison", index=False)
            ws5 = writer.sheets["5_PnL_Comparison"]
            style_header(ws5)
            hdr5 = [c.value for c in ws5[1]]
            autofit(ws5)
            # Colour P&L columns
            for col_name in ["manual_pnl_pct", "sys_pnl_pct"]:
                if col_name in hdr5:
                    ci = hdr5.index(col_name) + 1
                    for row in ws5.iter_rows(min_row=2, min_col=ci, max_col=ci):
                        for cell in row:
                            if isinstance(cell.value, (int, float)):
                                cell.fill = green_fill if cell.value > 0 else red_fill
            if "pnl_diff_pct" in hdr5:
                ci = hdr5.index("pnl_diff_pct") + 1
                for row in ws5.iter_rows(min_row=2, min_col=ci, max_col=ci):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.fill = green_fill if cell.value >= 0 else yellow_fill
            ms_col5 = hdr5.index("match_status") + 1 if "match_status" in hdr5 else None
            if ms_col5:
                colour_match_col(ws5, ms_col5)

            # Sheet 6: Missed Signals
            if not missed.empty:
                missed.to_excel(writer, sheet_name="6_Missed_Signals", index=False)
                ws6 = writer.sheets["6_Missed_Signals"]
                style_header(ws6)
                hdr6 = [c.value for c in ws6[1]]
                autofit(ws6, max_width=35)
                cl_idx6 = [i + 1 for i, h in enumerate(hdr6)
                           if h in list(CHECKLIST_RULES.keys())]
                colour_bool_cols(ws6, cl_idx6)

        print(f"\n  Audit report saved: {out_path}")
        logger.info("Audit saved: %s", out_path)

    except Exception as e:
        logger.error("Excel write failed: %s", e)
        csv_path = out_path.replace(".xlsx", ".csv")
        detail_df.to_csv(csv_path, index=False)
        print(f"  Saved (CSV fallback): {csv_path}")

    return detail_df


# ─── CLI entry point ───────────────────────────────────────────────────────
if __name__ == "__main__":
    import logging
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s  %(levelname)-8s  %(message)s",
                        datefmt="%H:%M:%S")
    run_audit(output_dir=OUTPUT_DIR)
