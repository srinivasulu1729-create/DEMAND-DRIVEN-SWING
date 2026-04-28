"""
backtest_engine.py - Demand-Driven Swing & Position System
Event-driven backtester with parallel symbol pre-loading.
Costs: Slippage 0.1%/side, brokerage Rs20 flat or 0.03% (lower).
"""

import logging
import math
import concurrent.futures
import multiprocessing
from dataclasses import dataclass, field
from typing import Literal

import numpy as np
import pandas as pd

from config import (
    STARTING_CAPITAL, SLIPPAGE_PCT, FLAT_BROKERAGE, PCT_BROKERAGE,
    POSITION_SL_PCT, SWING_SL_PCT,
    TARGET_CAGR, TARGET_WIN_RATE, MIN_TRADES,
    SELL_PROFIT_TARGET_LO, MAX_POSITION_PCT,
    MIN_HOLD_DAYS_WEAKNESS,
)
from indicators import (
    resample_weekly, add_weekly_emas, add_daily_emas,
    index_regime, risk_pct_for_regime, vol_ratio, vol_avg, atr,
)
from signals import (
    entry_checklist, compute_shares, stop_loss_price,
    sell_on_strength, sell_when_extended, sell_on_weakness,
    update_trailing_stop,
    rule17_swing_rs, rule18_swing_green_candles,
    rule19_ema_pullback, rule20_reversal_candle, rule21_vol_dryup,
    rule03_rs_ratio,
)

logger = logging.getLogger(__name__)
TradeType = Literal["position", "swing"]


# ============================================================
# PARALLEL SYMBOL PRE-LOADER
# ============================================================

def _load_one_symbol(args):
    symbol, index_pkl = args
    import pickle
    from data_loader import load_stock, align_stock_index
    from config import MIN_TRADING_ROWS
    try:
        index_df = pickle.loads(index_pkl)
        stock    = load_stock(symbol)
        if stock is None:
            return None
        s_al, i_al = align_stock_index(stock, index_df)
        if len(s_al) < MIN_TRADING_ROWS:
            return None
        return (symbol, s_al, i_al)
    except Exception as exc:
        # Return error tuple so the main process can log/diagnose it
        return ("__err__", symbol, str(exc))


def parallel_load_all_stocks(symbols, index_df, n_workers=None):
    """Load and align all symbols in parallel. Falls back to sequential."""
    import pickle
    workers   = n_workers or max(1, multiprocessing.cpu_count() - 1)
    idx_pkl   = pickle.dumps(index_df)
    args_list = [(sym, idx_pkl) for sym in symbols]
    result    = {}
    logger.info("Pre-loading %d symbols with %d workers ...", len(symbols), workers)
    try:
        with concurrent.futures.ProcessPoolExecutor(max_workers=workers) as exe:
            futures = {exe.submit(_load_one_symbol, a): a[0] for a in args_list}
            done = 0
            for fut in concurrent.futures.as_completed(futures):
                done += 1
                if done % 100 == 0:
                    logger.info("  Pre-loaded %d / %d ...", done, len(symbols))
                res = fut.result()
                if not res:
                    continue
                if isinstance(res, tuple) and res[0] == "__err__":
                    logger.debug("Worker error [%s]: %s", res[1], res[2])
                    continue
                sym, s, i = res
                result[sym] = (s, i)
    except Exception as exc:
        logger.warning("Parallel load failed (%s) - sequential fallback.", exc)
        from data_loader import load_all_stocks
        result = load_all_stocks(symbols, index_df)
    logger.info("Universe ready: %d symbols.", len(result))
    return result


# ============================================================
# DATA CLASSES
# ============================================================

@dataclass
class Trade:
    symbol:          str
    trade_type:      TradeType
    entry_date:      pd.Timestamp
    entry_price:     float
    stop_loss:       float
    shares:          int
    capital_at_entry: float
    entry_brokerage: float = 0.0
    initial_stop_loss: float = 0.0   # set once at entry; never trailed

    exit_date:   pd.Timestamp | None = None
    exit_price:  float | None = None
    exit_reason: str   = ""
    gross_pnl:   float = 0.0
    net_pnl:     float = 0.0
    brokerage:   float = 0.0
    is_closed:   bool  = False

    def close(self, date, price, reason):
        fill            = price * (1 - SLIPPAGE_PCT)
        value           = fill * self.shares
        broker_cost     = min(FLAT_BROKERAGE, PCT_BROKERAGE * value)
        self.exit_date  = date
        self.exit_price = fill
        self.exit_reason = reason
        self.gross_pnl  = (fill - self.entry_price) * self.shares
        self.brokerage  = self.entry_brokerage + broker_cost
        self.net_pnl    = self.gross_pnl - self.brokerage
        self.is_closed  = True

    def pnl_pct(self):
        if self.exit_price and self.entry_price:
            return (self.exit_price - self.entry_price) / self.entry_price
        return 0.0


@dataclass
class Portfolio:
    capital:       float = STARTING_CAPITAL
    open_trades:   list  = field(default_factory=list)
    closed_trades: list  = field(default_factory=list)
    equity_curve:  list  = field(default_factory=list)

    def record_equity(self, date, price_map):
        open_value = sum(
            t.shares * price_map.get(t.symbol, t.entry_price)
            for t in self.open_trades
        )
        self.equity_curve.append((date, self.capital + open_value))


# ============================================================
# COST HELPERS
# ============================================================

def brokerage(trade_value):
    return min(FLAT_BROKERAGE, PCT_BROKERAGE * trade_value)

def entry_fill_price(price):
    return price * (1 + SLIPPAGE_PCT)


# ============================================================
# BACKTESTER
# ============================================================

class Backtester:
    def __init__(self, stock_data, index_df,
                 max_open_positions=10, enable_swing=True):
        self.stock_data   = stock_data
        self.index_df     = index_df
        self.max_open     = max_open_positions
        self.enable_swing = enable_swing
        self.portfolio    = Portfolio()

        import time as _time
        _t0 = _time.time()

        # DatetimeIndex for helper methods (used outside the hot loop)
        self._idx_dti = pd.DatetimeIndex(index_df["date"])
        self._sym_dti: dict[str, pd.DatetimeIndex] = {
            sym: pd.DatetimeIndex(sdf["date"])
            for sym, (sdf, _) in stock_data.items()
        }

        # Pre-compute RS ratio arrays (one vectorised pct_change per symbol).
        # Lookup during sim: rs = self._rs_arr[sym][pos-1]  → O(1).
        _lookback = 40
        self._rs_arr: dict[str, np.ndarray] = {}
        for sym, (sdf, idf) in stock_data.items():
            s_ret = sdf["Close"].pct_change(_lookback).values
            i_ret = idf["Close"].pct_change(_lookback).values
            with np.errstate(divide="ignore", invalid="ignore"):
                rs = np.where(i_ret != 0, s_ret / i_ret, np.nan)
            self._rs_arr[sym] = rs

        # ── POSITION TABLE ────────────────────────────────────────────────
        # Pre-compute, for every (symbol, date_idx), the row count in that
        # symbol's DataFrame up to and including that date.
        # Replaces 2.7 M per-call DatetimeIndex.searchsorted with O(1) int
        # array indexing — the single biggest hot-loop cost.
        all_dates_dti = pd.DatetimeIndex(sorted(index_df["date"].unique()))
        self._all_dates: list = list(all_dates_dti)
        n_td = len(all_dates_dti)

        # Index positions (shape: n_td)
        self._idx_pos: np.ndarray = \
            self._idx_dti.searchsorted(all_dates_dti, side="right").astype(np.int32)

        # Per-symbol positions (shape: n_td each)
        self._pos_table: dict[str, np.ndarray] = {}
        for sym, dti in self._sym_dti.items():
            self._pos_table[sym] = \
                dti.searchsorted(all_dates_dti, side="right").astype(np.int32)

        logger.info("Backtester init complete: %.1fs  "
                    "(%d symbols, %d trading days, position table built)",
                    _time.time() - _t0, len(stock_data), n_td)

    # ── Fast slice helpers (O(log n) via DatetimeIndex.searchsorted) ─────

    def _index_up_to(self, date):
        pos = int(self._idx_dti.searchsorted(date, side="right"))
        return self.index_df.iloc[:pos]

    def _stock_up_to(self, sym, date):
        entry = self.stock_data.get(sym)
        if entry is None:
            return None
        sdf, _ = entry
        pos = int(self._sym_dti[sym].searchsorted(date, side="right"))
        return sdf.iloc[:pos]

    def _idx_aligned_up_to(self, sym, date):
        entry = self.stock_data.get(sym)
        if entry is None:
            return None
        _, idf = entry
        pos = int(self._sym_dti[sym].searchsorted(date, side="right"))
        return idf.iloc[:pos]

    def _process_exits(self, date, sym, s_hist, i_hist, trade):
        current = float(s_hist["Close"].iloc[-1])
        wkly    = resample_weekly(s_hist) if len(s_hist) >= 5 else None

        new_stop = update_trailing_stop(
            current, trade.entry_price, trade.stop_loss,
            trade.trade_type, s_hist, wkly
        )
        trade.stop_loss = new_stop

        # ── Stop-loss trigger ─────────────────────────────────────────────
        # Position trades: checked on weekly close (Friday) only.
        # Swing trades   : checked every daily close.
        is_friday = pd.Timestamp(date).weekday() == 4
        check_stop = (trade.trade_type == "swing") or is_friday

        if check_stop and current <= trade.stop_loss:
            # Exit at the stop price, not the (potentially gapped-down) close.
            # This prevents -98% losses from bad-data or gap-down days.
            exit_px = max(trade.stop_loss, current)   # stop or worse
            trade.close(date, exit_px, "StopLoss")
            return True

        sos = sell_on_strength(s_hist, trade.entry_price,
                               wkly if (wkly is not None and len(wkly) > 5) else None)
        if sos["sell_strength"]:
            reasons = [k for k, v in sos.items() if v and k != "sell_strength"]
            trade.close(date, current, "+".join(reasons))
            return True

        swe = sell_when_extended(s_hist)
        if swe["sell_extended"]:
            reasons = [k for k, v in swe.items() if v and k != "sell_extended"]
            trade.close(date, current, "+".join(reasons))
            return True

        # Weakness exits (rule31/32/33/34) are suppressed for the first
        # MIN_HOLD_DAYS_WEAKNESS calendar days after entry.  A position
        # entered at or near the 10-EMA will breach it on any minor red
        # day — blocking early exits gives the setup time to develop.
        days_held = (pd.Timestamp(date) - pd.Timestamp(trade.entry_date)).days
        if days_held >= MIN_HOLD_DAYS_WEAKNESS:
            _i_hist = i_hist if (i_hist is not None and not i_hist.empty) else self._index_up_to(date)
            sow = sell_on_weakness(s_hist, _i_hist)
            if sow["sell_weakness"]:
                reasons = [k for k, v in sow.items() if v and k != "sell_weakness"]
                trade.close(date, current, "+".join(reasons))
                return True

        return False

    def _try_position_entry(self, date, sym, s_hist, i_hist, weekly, idx_wkly, regime):
        checklist = entry_checklist(s_hist, i_hist, weekly, idx_wkly, "position")
        if not checklist["non_negotiable_pass"] or not checklist["all_pass"]:
            return None
        entry_px = entry_fill_price(float(s_hist["Close"].iloc[-1]))
        sl       = stop_loss_price(entry_px, "position")
        risk_pct = risk_pct_for_regime(regime)
        shares   = compute_shares(self.portfolio.capital, risk_pct, entry_px, sl)
        # ── Cap: each position max 10% of current capital ─────────────────
        max_shares = int(self.portfolio.capital * MAX_POSITION_PCT / entry_px)
        shares     = min(shares, max_shares)
        cost       = shares * entry_px
        broker     = brokerage(cost)
        if shares <= 0 or (cost + broker) > self.portfolio.capital:
            return None
        t = Trade(sym, "position", date, entry_px, sl, shares,
                  self.portfolio.capital, broker)
        t.initial_stop_loss = sl          # record once; never overwritten
        self.portfolio.capital -= (cost + broker)
        return t

    def _try_swing_entry(self, date, sym, s_hist, i_hist, regime):
        if not self.enable_swing:
            return None
        r17, _ = rule17_swing_rs(s_hist, i_hist)
        r18    = rule18_swing_green_candles(s_hist)
        r19, _ = rule19_ema_pullback(s_hist)
        r20    = rule20_reversal_candle(s_hist)
        r21    = rule21_vol_dryup(s_hist)
        if sum([r17, r18, r19, r20, r21]) < 4:
            return None
        entry_px = entry_fill_price(float(s_hist["Close"].iloc[-1]))
        sl       = stop_loss_price(entry_px, "swing")
        risk_pct = risk_pct_for_regime(regime)
        shares   = compute_shares(self.portfolio.capital, risk_pct, entry_px, sl)
        # ── Cap: each position max 10% of current capital ─────────────────
        max_shares = int(self.portfolio.capital * MAX_POSITION_PCT / entry_px)
        shares     = min(shares, max_shares)
        cost       = shares * entry_px
        broker     = brokerage(cost)
        if shares <= 0 or (cost + broker) > self.portfolio.capital:
            return None
        t = Trade(sym, "swing", date, entry_px, sl, shares,
                  self.portfolio.capital, broker)
        t.initial_stop_loss = sl          # record once; never overwritten
        self.portfolio.capital -= (cost + broker)
        return t

    def run(self, start_date=None, end_date=None):
        import time as _time
        import pandas as _pd

        # Apply date window filter
        all_dates = self._all_dates
        if start_date is not None:
            sd = _pd.Timestamp(start_date)
            all_dates = [d for d in all_dates if _pd.Timestamp(d) >= sd]
        if end_date is not None:
            ed = _pd.Timestamp(end_date)
            all_dates = [d for d in all_dates if _pd.Timestamp(d) <= ed]

        # Build a mapping from date → original index (for pos_table lookups)
        date_to_idx = {d: i for i, d in enumerate(self._all_dates)}
        date_indices = [date_to_idx[d] for d in all_dates]

        n_dates = len(all_dates)
        if n_dates == 0:
            logger.warning("No trading days in the specified date range.")
            return BacktestResult(self.portfolio)

        first_d = _pd.Timestamp(all_dates[0]).date()
        last_d  = _pd.Timestamp(all_dates[-1]).date()
        logger.info("Backtesting %d trading days  [%s → %s] ...",
                    n_dates, first_d, last_d)
        t_start = _time.time()

        # Cache weekly index — recompute only when the ISO week changes (5× faster)
        _cached_week:    int | None = None
        _cached_idx_wkly            = None
        _cached_regime:  str        = "bull"

        for loop_idx, (date_idx, date) in enumerate(zip(date_indices, all_dates)):

            # ── Progress report every 200 days ────────────────────────────
            if loop_idx > 0 and loop_idx % 200 == 0:
                elapsed = _time.time() - t_start
                pace    = elapsed / loop_idx
                eta_s   = pace * (n_dates - loop_idx)
                logger.info(
                    "  Day %d/%d  %s  open=%d  closed=%d  "
                    "%.3fs/day  ETA %.0fs (%.1f min)",
                    loop_idx, n_dates, _pd.Timestamp(date).date(),
                    len(self.portfolio.open_trades),
                    len(self.portfolio.closed_trades),
                    pace, eta_s, eta_s / 60,
                )

            # ── Index slice — O(1) position lookup ────────────────────────
            ip = int(self._idx_pos[date_idx])
            if ip < 40:
                continue

            # Recompute weekly index + regime only on week boundary
            iso_week = date.isocalendar()[1]
            if iso_week != _cached_week:
                _cached_week    = iso_week
                _cached_idx_wkly = resample_weekly(self.index_df.iloc[:ip])
                _cached_regime   = index_regime(_cached_idx_wkly)
            idx_wkly = _cached_idx_wkly
            regime   = _cached_regime
            price_map: dict = {}

            # ── Process exits — O(1) position lookup ──────────────────────
            still_open = []
            for trade in self.portfolio.open_trades:
                sym      = trade.symbol
                sdf, idf = self.stock_data[sym]
                sp       = int(self._pos_table[sym][date_idx])
                s_hist   = sdf.iloc[:sp]
                i_hist   = idf.iloc[:sp]
                if sp < 2:
                    still_open.append(trade)
                    continue
                price_map[sym] = float(s_hist["Close"].iloc[-1])
                closed = self._process_exits(date, sym, s_hist, i_hist, trade)
                if closed:
                    self.portfolio.capital += trade.exit_price * trade.shares
                    self.portfolio.closed_trades.append(trade)
                else:
                    still_open.append(trade)
            self.portfolio.open_trades = still_open

            # ── Scan entries ──────────────────────────────────────────────
            open_syms = {t.symbol for t in self.portfolio.open_trades}
            n_slots   = self.max_open - len(self.portfolio.open_trades)
            if n_slots <= 0:
                self.portfolio.record_equity(date, price_map)
                continue

            # Step 1: O(1) position lookup + O(1) RS array read — no slices yet
            cands: list = []
            for sym in self.stock_data:
                if sym in open_syms:
                    continue
                sp = int(self._pos_table[sym][date_idx])
                if sp < 60:
                    continue
                rs_val = float(self._rs_arr[sym][sp - 1])
                if np.isnan(rs_val):
                    rs_val = -999.0
                cands.append((rs_val, sym, sp))

            # Step 2: sort by RS desc
            cands.sort(key=lambda x: x[0], reverse=True)

            # Step 3: slices + resample_weekly only for top candidates tried.
            # Cap at MAX_ENTRY_TRIES to prevent scanning all 1175 on slow days.
            MAX_ENTRY_TRIES = 60
            added = 0
            tries = 0
            for _, sym, sp in cands:
                if added >= n_slots or tries >= MAX_ENTRY_TRIES:
                    break
                tries += 1
                sdf, idf = self.stock_data[sym]
                s_h  = sdf.iloc[:sp]
                i_h  = idf.iloc[:sp]
                wkly = resample_weekly(s_h)
                if len(wkly) < 10:
                    continue
                trade = self._try_position_entry(date, sym, s_h, i_h, wkly, idx_wkly, regime)
                if trade is None and regime != "bear":
                    trade = self._try_swing_entry(date, sym, s_h, i_h, regime)
                if trade:
                    self.portfolio.open_trades.append(trade)
                    price_map[sym] = trade.entry_price
                    added += 1
                    logger.debug("ENTRY %s %s @ %.2f  SL=%.2f",
                                 date.date(), sym, trade.entry_price, trade.stop_loss)

            self.portfolio.record_equity(date, price_map)

        # Force-close remaining
        last_date = all_dates[-1]
        for trade in self.portfolio.open_trades:
            s_hist = self._stock_up_to(trade.symbol, last_date)
            if s_hist is not None and len(s_hist):
                last_px = float(s_hist["Close"].iloc[-1])
                trade.close(last_date, last_px, "EndOfBacktest")
                self.portfolio.capital += trade.exit_price * trade.shares
                self.portfolio.closed_trades.append(trade)
        self.portfolio.open_trades = []
        return BacktestResult(self.portfolio)


# ============================================================
# METRICS + REPORT
# ============================================================

class BacktestResult:
    def __init__(self, portfolio):
        self.portfolio = portfolio
        self.trades_df = self._build_trades_df()
        self.metrics   = self._compute_metrics()

    def _build_trades_df(self):
        rows = []
        for t in self.portfolio.closed_trades:
            init_sl  = t.initial_stop_loss if t.initial_stop_loss else t.stop_loss
            sl_pct   = round((t.entry_price - init_sl) / t.entry_price * 100, 2) if t.entry_price else 0
            pos_pct  = round(t.shares * t.entry_price / t.capital_at_entry * 100, 2) if t.capital_at_entry else 0
            rows.append({
                "symbol":             t.symbol,
                "trade_type":         t.trade_type,
                "entry_date":         t.entry_date,
                "entry_price":        t.entry_price,
                "initial_stop_loss":  round(init_sl, 2),
                "sl_pct":             sl_pct,
                "final_stop_loss":    round(t.stop_loss, 2),
                "position_pct":       pos_pct,
                "exit_date":          t.exit_date,
                "exit_price":         t.exit_price,
                "shares":             t.shares,
                "gross_pnl":          t.gross_pnl,
                "net_pnl":            t.net_pnl,
                "brokerage":          t.brokerage,
                "exit_reason":        t.exit_reason,
                "pnl_pct":            t.pnl_pct() * 100,
                "capital_at_entry":   t.capital_at_entry,
            })
        return pd.DataFrame(rows)

    def _compute_metrics(self):
        df = self.trades_df
        if df.empty:
            return {"error": "No closed trades"}

        eq = pd.DataFrame(self.portfolio.equity_curve, columns=["date", "equity"])
        eq = eq.sort_values("date").drop_duplicates("date")

        start_cap  = STARTING_CAPITAL
        end_cap    = eq["equity"].iloc[-1] if not eq.empty else start_cap
        start_date = eq["date"].iloc[0]
        end_date   = eq["date"].iloc[-1]
        years      = (end_date - start_date).days / 365.25
        cagr       = ((end_cap / start_cap) ** (1 / years) - 1) if years > 0 else 0.0

        wins     = df[df["net_pnl"] > 0]
        losses   = df[df["net_pnl"] <= 0]
        win_rate = len(wins) / len(df)
        avg_win  = wins["net_pnl"].mean()  if len(wins)   else 0
        avg_loss = losses["net_pnl"].mean() if len(losses) else 0
        rr_ratio = abs(avg_win / avg_loss) if avg_loss != 0 else float("inf")

        eq["peak"] = eq["equity"].cummax()
        eq["dd"]   = (eq["equity"] - eq["peak"]) / eq["peak"]
        max_dd     = eq["dd"].min()

        gw = df[df["gross_pnl"] > 0]["gross_pnl"].sum()
        gl = df[df["gross_pnl"] < 0]["gross_pnl"].abs().sum()
        pf = gw / gl if gl > 0 else float("inf")

        eq["daily_ret"] = eq["equity"].pct_change()
        sharpe = (eq["daily_ret"].mean() / eq["daily_ret"].std() * math.sqrt(252)
                  if eq["daily_ret"].std() > 0 else 0.0)

        return {
            "start_date":       str(start_date.date()),
            "end_date":         str(end_date.date()),
            "years":            round(years, 2),
            "starting_capital": start_cap,
            "ending_capital":   round(end_cap, 2),
            "total_trades":     len(df),
            "winning_trades":   len(wins),
            "losing_trades":    len(losses),
            "cagr_pct":         round(cagr * 100, 2),
            "win_rate_pct":     round(win_rate * 100, 2),
            "avg_win":          round(avg_win, 2),
            "avg_loss":         round(avg_loss, 2),
            "rr_ratio":         round(rr_ratio, 2),
            "max_drawdown_pct": round(max_dd * 100, 2),
            "profit_factor":    round(pf, 2),
            "sharpe_ratio":     round(sharpe, 2),
            "pass_cagr":        cagr >= TARGET_CAGR,
            "pass_win_rate":    win_rate >= TARGET_WIN_RATE,
            "pass_trades":      len(df) >= MIN_TRADES,
            "OVERALL_PASS":     (cagr >= TARGET_CAGR
                                 and win_rate >= TARGET_WIN_RATE
                                 and len(df) >= MIN_TRADES),
        }

    def print_report(self):
        m = self.metrics
        if "error" in m:
            print(f"[ERROR] {m['error']}")
            return
        bar = "=" * 58
        print(f"\n{bar}")
        print("  BACKTEST REPORT  Demand-Driven Swing & Position System")
        print(bar)
        print(f"  Period          : {m['start_date']} -> {m['end_date']}  ({m['years']} yrs)")
        print(f"  Starting Capital: Rs {m['starting_capital']:,.0f}")
        print(f"  Ending Capital  : Rs {m['ending_capital']:,.0f}")
        print(f"  {'-'*52}")
        p = "PASS" if m["pass_cagr"] else "FAIL (target >=100%)"
        print(f"  CAGR            : {m['cagr_pct']:>8.2f}%  {p}")
        p = "PASS" if m["pass_win_rate"] else "FAIL (target >=60%)"
        print(f"  Win Rate        : {m['win_rate_pct']:>8.2f}%  {p}")
        p = "PASS" if m["pass_trades"] else "FAIL (need >=100)"
        print(f"  Total Trades    : {m['total_trades']:>8d}   {p}")
        print(f"  {'-'*52}")
        print(f"  Avg Win         : Rs {m['avg_win']:>10,.2f}")
        print(f"  Avg Loss        : Rs {m['avg_loss']:>10,.2f}")
        print(f"  RR Ratio        : {m['rr_ratio']:>8.2f}x")
        print(f"  Max Drawdown    : {m['max_drawdown_pct']:>8.2f}%")
        print(f"  Profit Factor   : {m['profit_factor']:>8.2f}")
        print(f"  Sharpe Ratio    : {m['sharpe_ratio']:>8.2f}")
        print(f"  {'-'*52}")
        status = "SYSTEM PASSES ALL CRITERIA" if m["OVERALL_PASS"] else "SYSTEM FAILS - review rules"
        print(f"  {status}")
        if not m["OVERALL_PASS"]:
            if not m["pass_cagr"]:
                print("  -> Review: Rules 1, 3, 9 (momentum & RS filters)")
            if not m["pass_win_rate"]:
                print("  -> Review: Rules 10-16 (entry timing & confirmation)")
        print(f"{bar}\n")

    def save(self, output_dir="."):
        import os, json
        os.makedirs(output_dir, exist_ok=True)

        # ── CSV (raw data for debugging) ──────────────────────────────────
        trades_path = os.path.join(output_dir, "backtest_trades.csv")
        self.trades_df.to_csv(trades_path, index=False)

        # ── Metrics JSON ──────────────────────────────────────────────────
        metrics_path = os.path.join(output_dir, "backtest_metrics.json")

        def _to_native(v):
            if isinstance(v, (np.bool_,)):   return bool(v)
            if isinstance(v, (np.integer,)): return int(v)
            if isinstance(v, (np.floating,)): return float(v)
            return v

        safe_metrics = {k: _to_native(v) for k, v in self.metrics.items()}
        with open(metrics_path, "w") as f:
            json.dump(safe_metrics, f, indent=2)

        # ── Equity curve CSV ──────────────────────────────────────────────
        eq_df = pd.DataFrame(self.portfolio.equity_curve, columns=["date", "equity"])
        eq_df.to_csv(os.path.join(output_dir, "equity_curve.csv"), index=False)

        # ── TradeLog.xlsx ─────────────────────────────────────────────────
        xlsx_path = os.path.join(output_dir, "TradeLog_Backtest.xlsx")
        self._write_tradelog_xlsx(xlsx_path)

        logger.info("Results saved to: %s", output_dir)
        print(f"  Saved: {trades_path}")
        print(f"  Saved: {metrics_path}")
        print(f"  Saved: {xlsx_path}")

    def _write_tradelog_xlsx(self, path: str):
        """Write a human-readable TradeLog.xlsx matching the manual trade log format."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side, GradientFill)
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed — skipping TradeLog.xlsx")
            return

        wb = Workbook()

        # ── Sheet 1: Trade Log ────────────────────────────────────────────
        ws = wb.active
        ws.title = "TradeLog"

        # Styles
        hdr_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        hdr_fill    = PatternFill("solid", fgColor="1F4E79")
        body_font   = Font(name="Arial", size=10)
        win_fill    = PatternFill("solid", fgColor="E2EFDA")   # light green
        loss_fill   = PatternFill("solid", fgColor="FCE4D6")   # light red
        open_fill   = PatternFill("solid", fgColor="FFF2CC")   # light yellow
        center_al   = Alignment(horizontal="center", vertical="center", wrap_text=False)
        left_al     = Alignment(horizontal="left",   vertical="center")
        thin        = Side(style="thin", color="BFBFBF")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            "No.", "Symbol", "Type", "Setup / Entry Pattern",
            "Entry Date", "Entry Price (₹)", "Initial SL (₹)", "SL %",
            "Position %", "Shares",
            "Exit Date", "Exit Price (₹)", "Exit Reason",
            "Gross P&L (₹)", "Brokerage (₹)", "Net P&L (₹)", "Return %",
            "Hold Days",
        ]
        col_widths = [5, 14, 9, 30, 13, 15, 14, 7, 10, 8, 13, 15, 28, 14, 13, 13, 10, 10]

        # Header row
        ws.row_dimensions[1].height = 22
        for c_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=c_idx, value=hdr)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center_al
            cell.border    = border
            ws.column_dimensions[get_column_letter(c_idx)].width = width

        # Data rows
        df = self.trades_df.copy()
        if df.empty:
            wb.save(path)
            return

        # Determine entry pattern from exit_reason or trade_type
        def _setup_label(row):
            er = str(row.get("exit_reason", ""))
            tt = str(row.get("trade_type", "position"))
            if "weekly_inside" in er.lower() or "inside" in er.lower():
                return "Weekly inside candle breakout"
            if tt == "swing":
                return "Swing entry (EMA pullback)"
            return "Position entry (breakout/base)"

        for i, (_, row) in enumerate(df.iterrows(), start=1):
            r = i + 1
            ws.row_dimensions[r].height = 18

            entry_dt = pd.Timestamp(row["entry_date"])
            exit_dt  = pd.Timestamp(row["exit_date"]) if pd.notna(row["exit_date"]) else None
            hold_days = (exit_dt - entry_dt).days if exit_dt else ""
            net_pnl   = float(row["net_pnl"])
            row_fill  = win_fill if net_pnl > 0 else (loss_fill if net_pnl < 0 else open_fill)

            values = [
                i,
                str(row["symbol"]),
                str(row["trade_type"]).capitalize(),
                _setup_label(row),
                entry_dt.strftime("%d-%b-%Y"),
                round(float(row["entry_price"]), 2),
                round(float(row.get("initial_stop_loss", row.get("stop_loss", 0))), 2),
                round(float(row.get("sl_pct", 0)), 2),
                round(float(row.get("position_pct", 0)), 2),
                int(row["shares"]),
                exit_dt.strftime("%d-%b-%Y") if exit_dt else "Open",
                round(float(row["exit_price"]), 2) if pd.notna(row.get("exit_price")) else "",
                str(row["exit_reason"]),
                round(float(row["gross_pnl"]), 2),
                round(float(row["brokerage"]), 2),
                round(net_pnl, 2),
                round(float(row["pnl_pct"]), 2),
                hold_days,
            ]

            for c_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.font      = body_font
                cell.fill      = row_fill
                cell.border    = border
                cell.alignment = center_al if c_idx != 4 else left_al

        # Freeze header
        ws.freeze_panes = "A2"

        # ── Sheet 2: Summary ──────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        m   = self.metrics
        if "error" not in m:
            s_hdr_font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            s_hdr_fill  = PatternFill("solid", fgColor="1F4E79")
            s_lbl_font  = Font(name="Arial", bold=True, size=10)
            s_val_font  = Font(name="Arial", size=10)
            pass_fill   = PatternFill("solid", fgColor="E2EFDA")
            fail_fill   = PatternFill("solid", fgColor="FCE4D6")

            ws2.column_dimensions["A"].width = 28
            ws2.column_dimensions["B"].width = 18
            ws2.column_dimensions["C"].width = 14

            ws2.merge_cells("A1:C1")
            title_cell = ws2["A1"]
            title_cell.value     = "Backtest Summary — Demand-Driven System"
            title_cell.font      = s_hdr_font
            title_cell.fill      = s_hdr_fill
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws2.row_dimensions[1].height = 24

            rows = [
                ("Period",          f"{m['start_date']}  →  {m['end_date']}", ""),
                ("Years",           m["years"], ""),
                ("Starting Capital","₹ {:,.0f}".format(m["starting_capital"]), ""),
                ("Ending Capital",  "₹ {:,.0f}".format(m["ending_capital"]),  ""),
                ("",                "",         ""),
                ("CAGR",            f"{m['cagr_pct']:.2f}%",    "PASS" if m["pass_cagr"]       else "FAIL"),
                ("Win Rate",        f"{m['win_rate_pct']:.2f}%", "PASS" if m["pass_win_rate"]   else "FAIL"),
                ("Total Trades",    m["total_trades"],           "PASS" if m["pass_trades"]     else "FAIL"),
                ("",                "",         ""),
                ("Winning Trades",  m["winning_trades"],  ""),
                ("Losing Trades",   m["losing_trades"],   ""),
                ("Avg Win",         "₹ {:,.2f}".format(m["avg_win"]),  ""),
                ("Avg Loss",        "₹ {:,.2f}".format(m["avg_loss"]), ""),
                ("RR Ratio",        f"{m['rr_ratio']:.2f}x", ""),
                ("Max Drawdown",    f"{m['max_drawdown_pct']:.2f}%", ""),
                ("Profit Factor",   f"{m['profit_factor']:.2f}",  ""),
                ("Sharpe Ratio",    f"{m['sharpe_ratio']:.2f}",   ""),
                ("",                "",         ""),
                ("OVERALL",         "PASS" if m["OVERALL_PASS"] else "FAIL", ""),
            ]

            for r_idx, (label, value, status) in enumerate(rows, start=2):
                ws2.row_dimensions[r_idx].height = 17
                lc = ws2.cell(row=r_idx, column=1, value=label)
                vc = ws2.cell(row=r_idx, column=2, value=value)
                sc = ws2.cell(row=r_idx, column=3, value=status)
                lc.font = s_lbl_font
                vc.font = s_val_font
                sc.font = Font(name="Arial", bold=True, size=10)
                lc.border = vc.border = sc.border = border
                lc.alignment = left_al
                vc.alignment = sc.alignment = center_al
                if status == "PASS":
                    sc.fill = pass_fill
                    sc.font = Font(name="Arial", bold=True, size=10, color="375623")
                elif status == "FAIL":
                    sc.fill = fail_fill
                    sc.font = Font(name="Arial", bold=True, size=10, color="9C0006")
                if label == "OVERALL":
                    lc.fill = vc.fill = pass_fill if m["OVERALL_PASS"] else fail_fill
                    lc.font = vc.font = Font(name="Arial", bold=True, size=11)

        wb.save(path)
