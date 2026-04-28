"""
live_screener.py — Demand-Driven Swing & Position System
Parallel daily screener using multiprocessing.Pool.
Section 8 output: Symbol, EntryType, Entry range, StopLoss,
Target, RS score, VolRatio, 10-item checklist.
"""

import logging
import os
import multiprocessing as mp
from datetime import datetime

import numpy as np
import pandas as pd

from config import (
    OUTPUT_DIR, SELL_PROFIT_TARGET_LO, SELL_PROFIT_TARGET_HI,
    POSITION_SL_PCT, SWING_SL_PCT,
)

logger = logging.getLogger(__name__)


# Worker must be at module level for pickle
def _screen_symbol(args):
    """Subprocess worker: screen one symbol. Returns dict or None."""
    import pickle
    symbol, idx_pkl, idxw_pkl = args
    index_df     = pickle.loads(idx_pkl)
    index_weekly = pickle.loads(idxw_pkl)
    try:
        from data_loader import load_stock, align_stock_index
        from indicators import resample_weekly, index_regime, vol_ratio
        from signals import (
            entry_checklist, rule03_rs_ratio,
            rule10_base_bottom_entry, rule14_new_high_breakout,
            rule19_ema_pullback,
        )
        stock_df = load_stock(symbol)
        if stock_df is None:
            return None
        stock_df, idx_aligned = align_stock_index(stock_df, index_df)
        if len(stock_df) < 60:
            return None
        weekly_df = resample_weekly(stock_df)
        if len(weekly_df) < 10:
            return None

        checklist = entry_checklist(stock_df, idx_aligned, weekly_df, index_weekly, "position")
        if not checklist["non_negotiable_pass"]:
            return None

        regime     = index_regime(index_weekly)
        current_px = float(stock_df["Close"].iloc[-1])
        _, rs_val  = rule03_rs_ratio(stock_df, idx_aligned)
        vr         = float(vol_ratio(stock_df).iloc[-1])

        if rule10_base_bottom_entry(stock_df):
            entry_type = "BaseBottom"
        elif rule14_new_high_breakout(stock_df):
            entry_type = "Breakout"
        else:
            ok, ema_tag = rule19_ema_pullback(stock_df)
            entry_type  = f"SwingEMA({ema_tag})" if ok else "Watchlist"

        if "Swing" in entry_type:
            t_lo = round(current_px * 1.05, 2)
            t_hi = round(current_px * 1.06, 2)
            sl   = round(current_px * (1 - SWING_SL_PCT), 2)
        else:
            t_lo = round(current_px * (1 + SELL_PROFIT_TARGET_LO), 2)
            t_hi = round(current_px * (1 + SELL_PROFIT_TARGET_HI), 2)
            sl   = round(current_px * (1 - POSITION_SL_PCT), 2)

        CL_KEYS = [
            "r01_price_50pct", "r02_green_weekly", "r03_rs_ratio",
            "r04_index_regime", "r_base_depth_8_25pct", "r11_vol_contracting",
            "r_entry_valid", "r_volume_confirms", "r_position_sized", "r08_rs_near_highs",
        ]
        n_pass = sum(1 for k in CL_KEYS if checklist.get(k, False))

        row = {
            "Symbol":         symbol,
            "EntryType":      entry_type,
            "EntryPriceLo":   round(current_px * 0.99, 2),
            "EntryPriceHi":   round(current_px * 1.01, 2),
            "StopLoss":       sl,
            "TargetLo":       t_lo,
            "TargetHi":       t_hi,
            "RS_Score":       round(rs_val, 2) if not np.isnan(rs_val) else None,
            "VolRatio":       round(vr, 2),
            "ChecklistScore": n_pass,
            "AllPass":        checklist["all_pass"],
            "CurrentPrice":   current_px,
            "Regime":         regime,
        }
        for k in CL_KEYS:
            row[f"CL_{k}"] = checklist.get(k, False)
        return row
    except Exception as exc:
        logger.debug("screen %s: %s", symbol, exc)
        return None


def screen(backtest_passed=True, n_workers=None):
    """Run parallel screener. Returns ranked DataFrame."""
    import pickle
    if not backtest_passed:
        logger.warning("Screener running WITHOUT confirmed backtest PASS.")

    from data_loader import load_index, load_symbols
    from indicators import resample_weekly, index_regime
    from signals import rule04_index_regime

    logger.info("Loading index ...")
    index_df     = load_index()
    index_weekly = resample_weekly(index_df)
    regime       = index_regime(index_weekly)

    if not rule04_index_regime(index_weekly):
        logger.warning("Index regime=%s — position trades BLOCKED (rule 4)", regime)

    symbols = load_symbols()
    logger.info("Screening %d symbols ...", len(symbols))

    idx_pkl  = pickle.dumps(index_df)
    idxw_pkl = pickle.dumps(index_weekly)
    args_list = [(sym, idx_pkl, idxw_pkl) for sym in symbols]
    workers   = n_workers or max(1, mp.cpu_count() - 1)
    logger.info("Using %d worker processes", workers)

    results = []
    with mp.Pool(processes=workers) as pool:
        for i, res in enumerate(pool.imap_unordered(_screen_symbol, args_list, chunksize=10), 1):
            if res is not None:
                results.append(res)
            if i % 200 == 0:
                logger.info("  %d / %d screened  (%d hits) ...", i, len(symbols), len(results))

    if not results:
        logger.warning("No stocks passed screening.")
        return pd.DataFrame()

    df = pd.DataFrame(results)
    df = df.sort_values(
        by=["AllPass", "ChecklistScore", "RS_Score"],
        ascending=[False, False, False],
    ).reset_index(drop=True)
    df.index = df.index + 1
    df.index.name = "Rank"
    logger.info("Screener: %d candidates  (%d AllPass)", len(df), int(df["AllPass"].sum()))
    return df


def print_screen_results(df, top_n=20):
    if df.empty:
        print("No results.")
        return
    print(f"\n{'=' * 100}")
    print(f"  LIVE SCREENER  {datetime.today().strftime('%d-%b-%Y')}  |  Top {top_n}")
    print(f"{'=' * 100}")
    cols = ["Symbol", "EntryType", "EntryPriceLo", "EntryPriceHi",
            "StopLoss", "TargetLo", "TargetHi", "RS_Score",
            "VolRatio", "ChecklistScore", "AllPass"]
    print(df[cols].head(top_n).to_string())
    print(f"{'=' * 100}")
    print(f"  AllPass: {int(df['AllPass'].sum())}  |  Total candidates: {len(df)}\n")


def save_screen_results(df, output_dir=OUTPUT_DIR):
    if df.empty:
        return
    stamp     = datetime.today().strftime("%Y%m%d")
    csv_path  = os.path.join(output_dir, f"screener_{stamp}.csv")
    xlsx_path = os.path.join(output_dir, f"screener_{stamp}.xlsx")
    df.to_csv(csv_path)
    try:
        from openpyxl.styles import PatternFill
        green = PatternFill("solid", fgColor="C6EFCE")
        red   = PatternFill("solid", fgColor="FFC7CE")

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            # Sheet 1: AllPass stocks only
            all_pass = df[df["AllPass"]].copy()
            all_pass.to_excel(writer, sheet_name="AllPass_Stocks")

            # Sheet 2: Full ranked list
            df.to_excel(writer, sheet_name="Full_Ranked_List")

            # Sheet 3: Checklist heatmap
            cl_cols = [c for c in df.columns if c.startswith("CL_")]
            if cl_cols:
                heat = df[["Symbol", "ChecklistScore"] + cl_cols].copy()
                heat.to_excel(writer, sheet_name="Checklist_Heatmap")
                ws3 = writer.sheets["Checklist_Heatmap"]
                hdr3 = [c.value for c in ws3[1]]
                for ci, col_name in enumerate(hdr3, 1):
                    if col_name and col_name.startswith("CL_"):
                        for row in ws3.iter_rows(min_row=2, min_col=ci, max_col=ci):
                            for cell in row:
                                cell.fill = green if cell.value else red

            # Colour AllPass column in full list
            ws2  = writer.sheets["Full_Ranked_List"]
            hdr2 = [c.value for c in ws2[1]]
            if "AllPass" in hdr2:
                ap = hdr2.index("AllPass") + 1
                for row in ws2.iter_rows(min_row=2, min_col=ap, max_col=ap):
                    for cell in row:
                        cell.fill = green if cell.value else red

        print(f"  Saved: {csv_path}")
        print(f"  Saved: {xlsx_path}")
    except Exception as e:
        logger.warning("xlsx save error: %s", e)
        print(f"  Saved: {csv_path}")
